"""
Conciliación BROU EUR - SAP vs Banco
Basado en conciliar_brou_uyu_v2.py

ESTRUCTURA SAP (EUROS.xlsx):
  - Header fila 0, cuenta fila 1, datos desde fila 2 (0-based)
  - Col 0: Fcha Contab | Col 2: Nº Doc | Col 6: Comentarios
  - Col 33 (0-based) = Sdo Vencido (ME) → importe EUR
  - Negativo = débito banco, positivo = crédito banco
  - Fila con ME=nan → diferencia de tipo de cambio, no tiene movimiento de caja → ROSA

ESTRUCTURA BANCO (xls/xlsx):
  - Headers fila 13 (0-based), datos desde fila 14
  - Col A=Fecha, B=Descripción, D=Nº doc, H=Débito, I=Crédito
  - Ignorar filas "Saldo inicial" / "Saldo final"
  - Débito = salida (negativo), Crédito = entrada (positivo)

COLORES (relleno SOLO en celda importe):
  Verde oscuro  = misma fecha + mismo importe EXACTO (diff == 0)
  Verde claro   = cierra con particularidad: diff centavos (0 < diff ≤ 0.9) O fecha ±1-3 días
  Amarillo      = comisiones bancarias (solo banco)
  Rosa          = no matchea / ME=nan
  Gris          = anulados SAP (par +/-)

REGLAS ESPECIALES:
  - Anulados: col Comentarios contiene "Cancelar" O "Anular" + "entrada para número de pago" → par en gris
  - Wiz: SAP líneas WizAAAAMMDDnX → agrupar/sumar vs línea única banco
  - Inverso Wiz: N líneas banco misma fecha+descripción cuya suma ≈ importe SAP
  - Tolerancia: abs(diff) <= 0.9 EUR
"""

import re
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# ── Colores ──────────────────────────────────────────────────────────────────
VERDE_CLARO  = PatternFill("solid", fgColor="92D051")
AMARILLO     = PatternFill("solid", fgColor="FFFF00")
ROSA         = PatternFill("solid", fgColor="FFB6C1")
GRIS         = PatternFill("solid", fgColor="C0C0C0")
SIN_RELLENO  = PatternFill(fill_type=None)

TOL = 0.9  # tolerancia EUR

# ── Helpers ───────────────────────────────────────────────────────────────────
def norm(s):
    """Normaliza unicode para comparaciones (ej: COMISIÓN → COMISION)."""
    return unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode().lower()

def parse_date(val):
    if pd.isna(val) or val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date() if hasattr(val, 'date') else val
    if hasattr(val, 'date'):
        return val.date()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except ValueError:
            pass
    return None

def importe_banco(row):
    """Devuelve importe neto: crédito positivo, débito negativo."""
    deb  = row['debito']
    cred = row['credito']
    if pd.notna(deb) and deb != 0:
        return -abs(float(deb))
    if pd.notna(cred) and cred != 0:
        return abs(float(cred))
    return 0.0

def es_comision(desc):
    n = norm(desc)
    return 'comisi' in n

def es_saldo(desc):
    n = norm(desc)
    return 'saldo inicial' in n or 'saldo final' in n

def wiz_code(doc, comentarios):
    """Extrae código Wiz si existe en doc o comentarios."""
    for s in [str(doc), str(comentarios)]:
        m = re.search(r'Wiz\d{8}n\d+', s, re.IGNORECASE)
        if m:
            return m.group(0)
    return None

def is_anulado_pair(df_sap):
    """
    Devuelve set de índices SAP que forman pares anulados → gris.

    Regla corregida:
    - Detecta tanto "Cancelar/Anular entrada para número de pago ..."
      como comentarios tipo "(Anulación) - 265201".
    - Busca el documento original por Nº Doc exacto normalizado.
    - Solo marca gris si el importe del original y el de anulación son opuestos.
      Ejemplo:
        AS 265201  EUR A UYU                   -30000  → gris
        AS 265222  EUR A UYU(Anulación)-265201 +30000  → gris
        AS 265223  EUR A UYU                   -30000  → NO gris, busca banco
    """
    gris_idx = set()

    cancel_pat = re.compile(
        r'(cancelar|anular|anulaci[oó]n).*?(entrada\s+para\s+n[uú]mero\s+de\s+pago)?',
        re.IGNORECASE
    )

    def only_digits(v):
        return re.sub(r'\D', '', str(v or ''))

    def imp_val(v):
        try:
            if v is None or pd.isna(v):
                return None
            return round(float(v), 2)
        except Exception:
            return None

    for i, row in df_sap.iterrows():
        com = str(row['comentarios'] or '')

        if not cancel_pat.search(com):
            continue

        # Prioridad 1: referencia luego de guion final: "... - 265201"
        m = re.search(r'[-–]\s*(\d{4,8})\s*$', com.strip())

        # Prioridad 2: número al final del comentario
        if not m:
            m = re.search(r'(\d{4,8})\s*$', com.strip())

        # Prioridad 3: número luego de "pago recibido" o "pago"
        if not m:
            m = re.search(r'pago(?:\s+recibido)?\s+(\d{4,8})', com, re.IGNORECASE)

        # Prioridad 4: último número largo del comentario
        ref = None
        if m:
            ref = m.group(1)
        else:
            nums = re.findall(r'\d{4,8}', com)
            if nums:
                ref = nums[-1]

        if not ref:
            continue

        imp_cancel = imp_val(row['importe'])
        if imp_cancel is None:
            continue

        # Buscar original exacto por Nº Doc normalizado y con importe opuesto.
        for j, row2 in df_sap.iterrows():
            if i == j:
                continue

            doc_digits = only_digits(row2['nro_doc'])
            if doc_digits != str(ref):
                continue

            imp_orig = imp_val(row2['importe'])
            if imp_orig is None:
                continue

            # Condición clave: deben cancelarse matemáticamente.
            if abs(imp_cancel + imp_orig) <= TOL:
                gris_idx.add(i)
                gris_idx.add(j)
                break

    return gris_idx


# ── Leer SAP ──────────────────────────────────────────────────────────────────
def leer_sap(path):
    df = pd.read_excel(path, header=None)
    # fila 0 = headers, fila 1 = totales cuenta, datos desde fila 2
    data_rows = []
    for i in range(2, len(df)):
        row = df.iloc[i]
        fecha = parse_date(row[0])
        nro_doc = str(row[2]) if pd.notna(row[2]) else ''
        comentarios = str(row[6]) if pd.notna(row[6]) else ''
        me_val = row[33]  # Sdo Vencido (ME)
        importe = None if pd.isna(me_val) else float(me_val)
        wiz = wiz_code(nro_doc, comentarios)
        data_rows.append({
            'excel_row': i + 1,   # 1-based para openpyxl
            'df_idx': i,
            'fecha': fecha,
            'nro_doc': nro_doc,
            'comentarios': comentarios,
            'importe': importe,
            'wiz': wiz,
            'matched': False,
            'color': None,
        })
    return pd.DataFrame(data_rows)

# ── Leer Banco ─────────────────────────────────────────────────────────────────
def leer_banco(path):
    df = pd.read_excel(path, header=None)
    # Headers en fila 13 (0-based), datos desde fila 14
    data_rows = []
    for i in range(14, len(df)):
        row = df.iloc[i]
        fecha = parse_date(row[0])
        desc  = str(row[1]) if pd.notna(row[1]) else ''
        nro   = str(row[3]) if pd.notna(row[3]) else ''
        deb   = row[7] if pd.notna(row[7]) else None
        cred  = row[8] if pd.notna(row[8]) else None

        if fecha is None and not desc.strip():
            continue
        if es_saldo(desc):
            continue

        data_rows.append({
            'excel_row': i + 1,
            'df_idx': i,
            'fecha': fecha,
            'descripcion': desc,
            'nro_doc': nro,
            'debito': deb,
            'credito': cred,
            'matched': False,
            'color': None,
        })
    df_b = pd.DataFrame(data_rows)
    df_b['importe'] = df_b.apply(importe_banco, axis=1)
    return df_b

# ── Matching ───────────────────────────────────────────────────────────────────
def _imp_null(val):
    """True si el importe es None o NaN (pandas convierte None a NaN en columnas numéricas)."""
    return val is None or (isinstance(val, float) and pd.isna(val))

def match_exacto(sap, banco, sap_idx, banco_used):
    """Mismo importe EXACTO (diff == 0) y misma fecha → Verde Oscuro."""
    row_s = sap.loc[sap_idx]
    if _imp_null(row_s['importe']):
        return None
    for bi, row_b in banco.iterrows():
        if bi in banco_used:
            continue
        if row_b['matched']:
            continue
        if row_s['importe'] == row_b['importe'] and row_s['fecha'] == row_b['fecha']:
            return bi
    return None

def match_claro(sap, banco, sap_idx, banco_used):
    """Cierra con particularidad → Verde Claro:
       - mismo importe con diff de centavos (0 < diff ≤ TOL) y misma fecha, O
       - mismo importe (o diff ≤ TOL) con fecha ±1-3 días."""
    row_s = sap.loc[sap_idx]
    if _imp_null(row_s['importe']):
        return None
    for bi, row_b in banco.iterrows():
        if bi in banco_used:
            continue
        if row_b['matched']:
            continue
        diff_imp = abs(row_s['importe'] - row_b['importe'])
        if diff_imp > TOL:
            continue
        if row_s['fecha'] and row_b['fecha']:
            diff_dias = abs((row_s['fecha'] - row_b['fecha']).days)
            # centavos misma fecha O cualquier diff dentro tolerancia ±3 días
            if diff_dias == 0 and diff_imp > 0:
                return bi
            if 1 <= diff_dias <= 3:
                return bi
    return None

def procesar_wiz(sap, banco, banco_used):
    """Agrupa líneas SAP por código Wiz, suma importes, cruza contra banco."""
    wiz_groups = {}
    for si, row_s in sap.iterrows():
        w = row_s['wiz']
        has_wiz = w is not None and not (isinstance(w, float) and pd.isna(w))
        if has_wiz:
            wiz_groups.setdefault(w, []).append(si)

    for wiz_code_val, sap_idxs in wiz_groups.items():
        total_sap = sum(sap.loc[i, 'importe'] for i in sap_idxs
                        if sap.loc[i, 'importe'] is not None)
        fecha_wiz = sap.loc[sap_idxs[0], 'fecha']

        # Buscar línea banco que coincida
        for bi, row_b in banco.iterrows():
            if bi in banco_used:
                continue
            if abs(total_sap - row_b['importe']) <= TOL:
                diff_dias = abs((fecha_wiz - row_b['fecha']).days) if fecha_wiz and row_b['fecha'] else 99
                color = VERDE_CLARO if diff_dias == 0 else VERDE_CLARO if diff_dias <= 3 else None
                if color:
                    banco_used.add(bi)
                    banco.at[bi, 'matched'] = True
                    banco.at[bi, 'color'] = color
                    for si in sap_idxs:
                        sap.at[si, 'matched'] = True
                        sap.at[si, 'color'] = color
                    break

def procesar_inverso_wiz(sap, banco, banco_used):
    """N líneas banco misma fecha+descripción cuya suma ≈ importe SAP."""
    for si, row_s in sap.iterrows():
        if row_s['matched']:
            continue
        if _imp_null(row_s['importe']):
            continue
        grupos = {}
        for bi, row_b in banco.iterrows():
            if bi in banco_used:
                continue
            if row_b['fecha'] != row_s['fecha']:
                continue
            key = (row_b['fecha'], norm(row_b['descripcion']))
            grupos.setdefault(key, []).append(bi)

        for key, bis in grupos.items():
            total_banco = sum(banco.loc[bi, 'importe'] for bi in bis)
            if abs(row_s['importe'] - total_banco) <= TOL:
                color = VERDE_CLARO  # misma fecha
                sap.at[si, 'matched'] = True
                sap.at[si, 'color'] = color
                for bi in bis:
                    banco_used.add(bi)
                    banco.at[bi, 'matched'] = True
                    banco.at[bi, 'color'] = color
                break

# ── Aplicar colores ────────────────────────────────────────────────────────────
def aplicar_color_sap(ws, excel_row, color):
    """Aplica relleno en columna D (importe SAP = col 34 openpyxl = col AH)."""
    # SAP importe está en columna 34 (1-based) = col AH
    col_importe = 34  # col 33 0-based + 1 = 34 openpyxl
    cell = ws.cell(row=excel_row, column=col_importe)
    cell.fill = color

def aplicar_color_banco(ws, excel_row, color):
    """Aplica relleno en columna H (débito) o I (crédito) — solo donde hay valor."""
    # Débito col H=8, Crédito col I=9 (1-based openpyxl)
    for col in [8, 9]:
        cell = ws.cell(row=excel_row, column=col)
        if cell.value is not None and str(cell.value).strip() not in ('', 'None'):
            cell.fill = color

# ── Main ───────────────────────────────────────────────────────────────────────
def conciliar(path_sap, path_banco, path_banco_orig_xls):
    print("Leyendo SAP...")
    sap = leer_sap(path_sap)
    print(f"  {len(sap)} movimientos SAP")

    print("Leyendo Banco...")
    banco = leer_banco(path_banco)
    print(f"  {len(banco)} movimientos Banco")

    banco_used = set()

    # 1. Marcar anulados SAP en gris
    gris_idx = is_anulado_pair(sap)
    for si in gris_idx:
        sap.at[si, 'matched'] = True
        sap.at[si, 'color'] = GRIS

    # 1b. Marcar diferencias de tipo de cambio en gris (ME=nan, no son movimientos de caja)
    for si, row_s in sap.iterrows():
        if row_s['matched']:
            continue
        if row_s['importe'] is None or (isinstance(row_s['importe'], float) and pd.isna(row_s['importe'])):
            sap.at[si, 'matched'] = True
            sap.at[si, 'color'] = GRIS

    # 2. Marcar comisiones banco en amarillo
    for bi, row_b in banco.iterrows():
        if es_comision(row_b['descripcion']):
            banco.at[bi, 'matched'] = True
            banco.at[bi, 'color'] = AMARILLO
            banco_used.add(bi)

    # 3. Procesar Wiz
    procesar_wiz(sap, banco, banco_used)

    # 4. Inverso Wiz
    procesar_inverso_wiz(sap, banco, banco_used)

    # 5. Match exacto (fecha + importe idéntico)
    for si, row_s in sap.iterrows():
        if row_s['matched']:
            continue
        if _imp_null(row_s['importe']):
            continue
        bi = match_exacto(sap, banco, si, banco_used)
        if bi is not None:
            sap.at[si, 'matched'] = True
            sap.at[si, 'color'] = VERDE_CLARO
            banco.at[bi, 'matched'] = True
            banco.at[bi, 'color'] = VERDE_CLARO
            banco_used.add(bi)

    # 6. Match verde claro (centavos o ±3 días)
    for si, row_s in sap.iterrows():
        if row_s['matched']:
            continue
        if _imp_null(row_s['importe']):
            continue
        bi = match_claro(sap, banco, si, banco_used)
        if bi is not None:
            sap.at[si, 'matched'] = True
            sap.at[si, 'color'] = VERDE_CLARO
            banco.at[bi, 'matched'] = True
            banco.at[bi, 'color'] = VERDE_CLARO
            banco_used.add(bi)

    # 7. No matcheados → rosa
    for si, row_s in sap.iterrows():
        if not row_s['matched']:
            sap.at[si, 'color'] = ROSA

    for bi, row_b in banco.iterrows():
        if not row_b['matched']:
            banco.at[bi, 'color'] = ROSA

    # ── Aplicar colores a SAP xlsx y limpiar columnas irrelevantes ───────────
    print("Aplicando colores SAP...")
    wb_sap = load_workbook(path_sap)
    ws_sap = wb_sap.active

    # Columnas a conservar (1-based openpyxl): 1=Fcha Contab, 3=Nº Doc, 7=Comentarios, 34=Sdo Vencido ME
    COLS_KEEP = {1, 3, 7, 34}
    total_cols = ws_sap.max_column

    # Aplicar colores ANTES de eliminar columnas (índices todavía son los originales)
    for _, row_s in sap.iterrows():
        if row_s['color'] is not None:
            aplicar_color_sap(ws_sap, row_s['excel_row'], row_s['color'])

    # Eliminar columnas no relevantes (de derecha a izquierda para no desplazar índices)
    cols_to_delete = sorted([c for c in range(1, total_cols + 1) if c not in COLS_KEEP], reverse=True)
    for col_idx in cols_to_delete:
        ws_sap.delete_cols(col_idx)

    # Después de eliminar, el importe quedó en la 4ta columna (D)
    # Ajustar ancho de columnas
    ws_sap.column_dimensions['A'].width = 14
    ws_sap.column_dimensions['B'].width = 14
    ws_sap.column_dimensions['C'].width = 50
    ws_sap.column_dimensions['D'].width = 18

    out_sap = 'SAP_EUR_conciliado.xlsx'
    wb_sap.save(out_sap)
    print(f"  Guardado: {out_sap}")

    # ── Aplicar colores a BANCO xlsx ──────────────────────────────────────────
    print("Aplicando colores Banco...")
    wb_banco = load_workbook(path_banco)
    ws_banco = wb_banco.active

    comisiones = []
    for _, row_b in banco.iterrows():
        color = row_b['color']
        if color is not None:
            aplicar_color_banco(ws_banco, row_b['excel_row'], color)
        if row_b['color'] is AMARILLO or (isinstance(row_b['color'], PatternFill) and row_b['color'].fgColor.rgb == 'FFFFFF00'):
            comisiones.append(row_b)

    # Fila de total comisiones al pie (en amarillo)
    total_com = sum(abs(r['importe']) for r in comisiones)
    if total_com > 0:
        # Encontrar última fila con datos
        last_row = ws_banco.max_row + 2
        ws_banco.cell(row=last_row, column=1).value = 'TOTAL COMISIONES'
        ws_banco.cell(row=last_row, column=8).value = total_com
        ws_banco.cell(row=last_row, column=1).fill = AMARILLO
        ws_banco.cell(row=last_row, column=8).fill = AMARILLO

    out_banco = 'BANCO_EUR_conciliado.xlsx'
    wb_banco.save(out_banco)
    print(f"  Guardado: {out_banco}")

    # ── Resumen ───────────────────────────────────────────────────────────────
    sap_total = len(sap)
    sap_matcheados = sap['matched'].sum()
    banco_total = len(banco)
    banco_matcheados = banco['matched'].sum()
    banco_comisiones = (banco['color'].apply(lambda c: c is not None and isinstance(c, PatternFill) and c.fgColor.rgb == 'FFFFFF00')).sum()

    print("\n" + "="*50)
    print("RESUMEN CONCILIACIÓN BROU EUR")
    print("="*50)
    print(f"SAP:   {sap_matcheados}/{sap_total} matcheados")
    print(f"Banco: {banco_matcheados}/{banco_total} matcheados")
    if total_com > 0:
        print(f"Comisiones: {total_com:.2f} EUR")

    print("\nSAP NO MATCHEADOS:")
    for _, r in sap[~sap['matched']].iterrows():
        print(f"  Row {r['excel_row']}: {r['fecha']} | {r['nro_doc']} | {r['importe']} | {r['comentarios'][:50]}")

    print("\nBANCO NO MATCHEADOS:")
    for _, r in banco[~banco['matched']].iterrows():
        print(f"  Row {r['excel_row']}: {r['fecha']} | {r['descripcion'][:40]} | {r['importe']}")

    return out_sap, out_banco

if __name__ == '__main__':
    conciliar('sap_eur.xlsx', 'banco_eur.xlsx', 'banco_eur.xls')
