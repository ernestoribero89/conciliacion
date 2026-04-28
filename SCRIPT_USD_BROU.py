"""
CONCILIACIÓN BANCARIA - BROU USD
v2.0 — reglas actualizadas

USO:
    python conciliar_brou_usd.py <archivo_sap.xlsx> <archivo_banco.xls/xlsx>

SALIDA:
    CONCILIACION_SAP.xlsx   — SAP con 4 columnas + color en celda importe
    CONCILIACION_BANCO.xlsx — Banco original + color en celda importe + total comisiones al pie

COLORES (relleno solo en celda importe, nunca en otras celdas):
    Verde claro  = match exacto (mismo importe ±0.9, fecha ±3 días)
    Verde oscuro = cierra matemáticamente pero con particularidad
                   (Wiz combinado, anulado+reemisión, combinación inusual)
    Amarillo     = comisión bancaria
    Rosa         = sin match (revisión manual)
    Gris         = anulados SAP + diferencias de tipo de cambio (excluidos del cruce)

COLUMNA IMPORTE SAP:
    BROU USD → 'Sdo vencido USD'  (col 16, 0-based)
    BROU UYU/EUR → 'Sdo Vencido (ME)' — usar el script UYU para esos casos
"""

import sys, re, unicodedata
import pandas as pd
from itertools import combinations
from datetime import timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import shutil
from collections import defaultdict

# ── Colores ───────────────────────────────────────────────────────────────────
VERDE_CLARO  = PatternFill('solid', fgColor='92D051')
AMARILLO     = PatternFill('solid', fgColor='FFFF00')
ROSA         = PatternFill('solid', fgColor='FF9999')
GRIS         = PatternFill('solid', fgColor='BFBFBF')

TOL      = 0.9
DATE_FMT = 'DD/MM/YYYY'

wiz_re    = re.compile(r'(Wiz\d{8}n\d+)', re.IGNORECASE)
cancel_re = re.compile(
    r'(?:cancelar|anular)\s+entrada\s+para\s+n[uú]mero\s+de\s+pago(?:\s+recibido)?\s+(\S+)',
    re.IGNORECASE
)

def norm(s):
    return unicodedata.normalize('NFD', str(s)).encode('ascii', 'ignore').decode('ascii').lower()

def get_wiz(c):
    m = wiz_re.search(str(c))
    return m.group(1) if m else None

def has_wiz(v):
    return v is not None and not (isinstance(v, float) and pd.isna(v))

def is_comision(d):
    n = norm(d)
    return 'comision' in n or 'apertura coe' in n

def is_dif_cambio(c):
    cl = norm(c)
    return ('diferencia' in cl or ' dif' in cl) and ('cambio' in cl or 'tipo' in cl)

def wiz_fecha(w):
    m = re.search(r'Wiz(\d{4})(\d{2})(\d{2})', w, re.IGNORECASE)
    if m: return pd.Timestamp(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    return None


# ── Carga SAP ─────────────────────────────────────────────────────────────────
def cargar_sap(path):
    """
    Fila 0 = headers, fila 1 = subtítulo (Activos), datos desde fila 2.
    Importe: columna 'Sdo vencido USD' (col 16, 0-based).
    Negativo = débito banco, positivo = crédito banco.
    """
    raw = pd.read_excel(path, header=None)
    headers = raw.iloc[0].tolist()
    col_fecha   = headers.index('Fcha Contab.')
    col_ndoc    = headers.index('Nº Doc')
    col_com     = headers.index('Comentarios')
    col_importe = headers.index('Sdo vencido USD')  # col 16 (0-based)

    df = raw.iloc[2:, [col_fecha, col_ndoc, col_com, col_importe]].copy().reset_index(drop=False)
    df.columns = ['orig_idx', 'fecha', 'ndoc', 'comentarios', 'importe']
    df['fecha']   = pd.to_datetime(df['fecha'], dayfirst=True, errors='coerce')
    df['importe'] = pd.to_numeric(df['importe'], errors='coerce')
    df = df.dropna(subset=['fecha', 'importe'])
    df['wiz']          = df['comentarios'].apply(get_wiz)
    df['es_dif_cambio'] = df['comentarios'].apply(is_dif_cambio)
    return df


# ── Carga BANCO ───────────────────────────────────────────────────────────────
def cargar_banco(path):
    """
    BROU USD: datos desde fila 3 (0-based idx=2), sin header row explícita.
    Cols: A=Fecha, B=Descripcion, C=vacía, D=Nº doc, E=vacía,
          F=Asunto, G=Dependencia, H=Débito, I=Crédito
    Signo: crédito=positivo, débito=negativo.
    Excluye filas de Saldo.
    """
    raw = pd.read_excel(path, header=None)
    df  = raw.iloc[2:].copy().reset_index(drop=False)
    df.columns = ['orig_idx','fecha','descripcion','_c','ndoc','_e',
                  'asunto','dependencia','debito','credito']
    df['fecha']   = pd.to_datetime(df['fecha'], dayfirst=True, errors='coerce')
    df['debito']  = pd.to_numeric(df['debito'],  errors='coerce').fillna(0)
    df['credito'] = pd.to_numeric(df['credito'], errors='coerce').fillna(0)
    df['importe'] = df['credito'] - df['debito']
    df = df.dropna(subset=['fecha'])
    df = df[~df['descripcion'].astype(str).str.lower().str.contains('saldo', na=False)]
    df['es_comision'] = df['descripcion'].apply(is_comision)
    return df


# ── Motor de conciliación ─────────────────────────────────────────────────────
def conciliar(sap, banco):
    sap_col   = {}   # orig_idx -> PatternFill
    bco_col   = {}
    used_bco  = set()

    # ── 1. Anulados + diferencias de cambio → GRIS ────────────────────────────
    anulados = set()
    for _, row in sap.iterrows():
        if row['es_dif_cambio']:
            anulados.add(row['orig_idx'])
            continue
        m = cancel_re.search(str(row['comentarios']))
        if m:
            anulados.add(row['orig_idx'])
            doc = m.group(1).strip()
            for oidx in sap[sap['ndoc'].astype(str).str.contains(doc, na=False)]['orig_idx']:
                anulados.add(oidx)

    for oidx in anulados:
        sap_col[oidx] = GRIS

    # ── 2. Wiz individual: N SAP → 1 banco ───────────────────────────────────
    wiz_groups = defaultdict(list)
    for _, row in sap.iterrows():
        if row['orig_idx'] in anulados: continue
        if has_wiz(row['wiz']):
            wiz_groups[row['wiz']].append(row)

    wiz_matched = set()
    for wiz_code, rows in wiz_groups.items():
        total = round(sum(r['importe'] for r in rows), 2)
        fs    = wiz_fecha(wiz_code) or rows[0]['fecha']
        for _, brow in banco.iterrows():
            if brow['orig_idx'] in used_bco: continue
            if abs(abs(round(brow['importe'], 2)) - abs(total)) <= TOL:
                if abs((brow['fecha'] - fs).days) <= 3:
                    for r in rows: sap_col[r['orig_idx']] = VERDE_CLARO
                    bco_col[brow['orig_idx']] = VERDE_CLARO
                    used_bco.add(brow['orig_idx'])
                    wiz_matched.add(wiz_code)
                    break

    # ── 3. Wiz combinado: múltiples grupos sin match misma fecha → 1 banco ───
    wiz_sin_match = {w: rows for w, rows in wiz_groups.items() if w not in wiz_matched}
    by_fecha_wiz = defaultdict(list)
    for w, rows in wiz_sin_match.items():
        fd = wiz_fecha(w)
        if fd: by_fecha_wiz[fd].append((w, rows))

    for fd, grupo in by_fecha_wiz.items():
        all_rows = [r for _, rows in grupo for r in rows if r['orig_idx'] not in sap_col]
        if not all_rows: continue
        total = round(sum(r['importe'] for r in all_rows), 2)
        for _, brow in banco.iterrows():
            if brow['orig_idx'] in used_bco: continue
            if abs(abs(round(brow['importe'], 2)) - abs(total)) <= TOL:
                if abs((brow['fecha'] - fd).days) <= 3:
                    for r in all_rows:
                        sap_col[r['orig_idx']] = VERDE_CLARO
                    bco_col[brow['orig_idx']] = VERDE_CLARO
                    used_bco.add(brow['orig_idx'])
                    break

    # ── 4. Wiz + anulado + reemisión → VERDE OSCURO ──────────────────────────
    # Un PP del grupo Wiz fue anulado; el banco pagó (Wiz - anulado + reemisión).
    for wiz_code, rows in wiz_groups.items():
        if all(r['orig_idx'] in sap_col for r in rows): continue
        fd = wiz_fecha(wiz_code) or rows[0]['fecha']
        docs_grupo = {re.sub(r'\D', '', str(r['ndoc'])) for r in rows}

        # Buscar filas de cancelación que referencien docs de este grupo
        anulados_ref = []
        for _, row_c in sap.iterrows():
            if row_c['orig_idx'] not in anulados: continue
            m2 = re.search(r'(\d{4,6})\s*$', str(row_c['comentarios']).strip())
            if not m2: continue
            ref = m2.group(1)
            if ref in docs_grupo:
                orig = next((r for r in rows if re.sub(r'\D','',str(r['ndoc'])) == ref), None)
                if orig: anulados_ref.append((row_c, orig))

        if not anulados_ref: continue

        orig_anulados_idx = {orig['orig_idx'] for _, orig in anulados_ref}
        total_neto = round(sum(r['importe'] for r in rows if r['orig_idx'] not in orig_anulados_idx), 2)

        # PP sueltos (sin Wiz, no anulados, no matcheados) cerca de fd
        sueltos = [(oidx, row) for oidx, row in
                   [(r['orig_idx'], r) for _, r in sap.iterrows()]
                   if oidx not in sap_col and not has_wiz(sap.loc[sap['orig_idx']==oidx, 'wiz'].values[0] if len(sap.loc[sap['orig_idx']==oidx])>0 else None)
                   and abs((sap.loc[sap['orig_idx']==oidx, 'fecha'].values[0] - fd).days) <= 3]

        # Más simple: iterar sap directo
        for _, row_s in sap.iterrows():
            if row_s['orig_idx'] in sap_col: continue
            if has_wiz(row_s['wiz']): continue
            if abs((row_s['fecha'] - fd).days) > 3: continue
            total_real = round(total_neto + row_s['importe'], 2)
            for _, brow in banco.iterrows():
                if brow['orig_idx'] in used_bco: continue
                if abs(abs(round(brow['importe'], 2)) - abs(total_real)) <= TOL:
                    if abs((brow['fecha'] - fd).days) <= 3:
                        for r in rows:
                            sap_col[r['orig_idx']] = VERDE_CLARO
                        for row_c, _ in anulados_ref:
                            sap_col[row_c['orig_idx']] = VERDE_CLARO
                        sap_col[row_s['orig_idx']] = VERDE_CLARO
                        bco_col[brow['orig_idx']] = VERDE_CLARO
                        used_bco.add(brow['orig_idx'])
                        break
            else:
                continue
            break

    # ── 5. Inverso Wiz: 1 SAP → N banco misma fecha+desc ─────────────────────
    bco_groups = defaultdict(list)
    for _, brow in banco.iterrows():
        if brow['es_comision']: continue
        key = (brow['fecha'].date(), str(brow['descripcion']).strip())
        bco_groups[key].append(brow)

    for _, row in sap.iterrows():
        oidx = row['orig_idx']
        if oidx in sap_col: continue
        if has_wiz(row['wiz']): continue
        sap_imp  = round(row['importe'], 2)
        sap_date = row['fecha'].date()

        best_color = None; best_rows = None; best_delta = 999
        for delta in range(4):
            for sign in ([0] if delta == 0 else [1, -1]):
                check = sap_date + timedelta(days=delta * sign)
                for (bd, bdesc), brows in bco_groups.items():
                    if bd != check: continue
                    avail = [r for r in brows if r['orig_idx'] not in used_bco]
                    if len(avail) < 2: continue
                    for n in range(2, len(avail) + 1):
                        for combo in combinations(avail, n):
                            total = round(sum(r['importe'] for r in combo), 2)
                            if abs(abs(total) - abs(sap_imp)) <= TOL:
                                d = abs((check - sap_date).days)
                                if d < best_delta:
                                    best_delta = d
                                    best_rows  = combo
                                    best_color = VERDE_CLARO

        if best_rows is not None:
            sap_col[oidx] = best_color
            for br in best_rows:
                bco_col[br['orig_idx']] = best_color
                used_bco.add(br['orig_idx'])

    # ── 6. Individual 1:1 ────────────────────────────────────────────────────
    for _, row in sap.iterrows():
        oidx = row['orig_idx']
        if oidx in sap_col: continue
        if has_wiz(row['wiz']): continue
        sap_imp  = round(row['importe'], 2)
        sap_date = row['fecha'].date()
        best = None; best_delta = 999
        for _, brow in banco.iterrows():
            if brow['orig_idx'] in used_bco: continue
            if brow['es_comision']: continue
            bimp = round(brow['importe'], 2)
            if abs(abs(bimp) - abs(sap_imp)) <= TOL:
                delta = abs((brow['fecha'].date() - sap_date).days)
                if delta <= 3 and delta < best_delta:
                    best = brow; best_delta = delta
        if best is not None:
            sap_col[oidx] = VERDE_CLARO
            bco_col[best['orig_idx']] = VERDE_CLARO
            used_bco.add(best['orig_idx'])

    # ── 7. Comisiones → AMARILLO ─────────────────────────────────────────────
    for _, brow in banco.iterrows():
        if brow['es_comision'] and brow['orig_idx'] not in used_bco:
            bco_col[brow['orig_idx']] = AMARILLO
            used_bco.add(brow['orig_idx'])

    # ── 8. Resto → ROSA ──────────────────────────────────────────────────────
    for _, row in sap.iterrows():
        if row['orig_idx'] not in sap_col:
            sap_col[row['orig_idx']] = ROSA
    for _, brow in banco.iterrows():
        if brow['orig_idx'] not in bco_col:
            bco_col[brow['orig_idx']] = ROSA

    return sap_col, bco_col


# ── Escritura SAP ─────────────────────────────────────────────────────────────
def escribir_sap(sap, sap_col, path_out):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Conciliacion'
    ws.append(['Fcha Contab.', 'Nº Doc', 'Comentarios', 'Sdo vencido USD'])
    ws.append(['Activos', '', 'Brou C/C U$S', ''])
    for _, row in sap.iterrows():
        ws.append([row['fecha'], row['ndoc'], row['comentarios'], row['importe']])
        er = ws.max_row
        ws.cell(row=er, column=4).fill = sap_col.get(row['orig_idx'], ROSA)
        ws.cell(row=er, column=1).number_format = DATE_FMT
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 16
    wb.save(path_out)


# ── Escritura BANCO ───────────────────────────────────────────────────────────
def escribir_banco(banco, bco_col, path_in, path_out):
    shutil.copy(path_in, path_out)
    wb = load_workbook(path_out)
    ws = wb.active
    for _, brow in banco.iterrows():
        oidx = brow['orig_idx']
        fill = bco_col.get(oidx)
        if fill:
            col = 9 if brow['credito'] > 0 else 8
            ws.cell(row=oidx + 1, column=col).fill = fill
        cell = ws.cell(row=oidx + 1, column=1)
        if cell.value is not None:
            cell.number_format = DATE_FMT

    coms = banco[banco['orig_idx'].isin([k for k, v in bco_col.items() if v == AMARILLO])]
    if len(coms) > 0:
        total_com = round(coms['debito'].sum(), 2)
        last_row  = max(banco['orig_idx']) + 2
        ws.cell(row=last_row, column=1, value='TOTAL COMISIONES').fill = AMARILLO
        ws.cell(row=last_row, column=8, value=total_com).fill = AMARILLO

    wb.save(path_out)


# ── Resumen consola ───────────────────────────────────────────────────────────
def imprimir_resumen(sap, banco, sap_col, bco_col):
    verdes = (VERDE_CLARO,)
    sv = sum(1 for c in sap_col.values() if c in verdes)
    sr = sum(1 for c in sap_col.values() if c == ROSA)
    sg = sum(1 for c in sap_col.values() if c == GRIS)
    bv = sum(1 for c in bco_col.values() if c in verdes)
    ba = sum(1 for c in bco_col.values() if c == AMARILLO)
    br = sum(1 for c in bco_col.values() if c == ROSA)
    coms = banco[banco['orig_idx'].isin([k for k,v in bco_col.items() if v == AMARILLO])]
    total_com = round(coms['debito'].sum(), 2)

    print('─' * 52)
    print(f'SAP   total={len(sap):>4}  verde={sv:>4}  rosa={sr:>3}  gris={sg:>3}')
    print(f'BANCO total={len(banco):>4}  verde={bv:>4}  amarillo={ba:>3}  rosa={br:>3}')
    print(f'Total comisiones: {total_com:,.2f} USD')
    print('─' * 52)

    if sr > 0:
        print('\nSAP sin match:')
        for _, row in sap.iterrows():
            if sap_col.get(row['orig_idx']) == ROSA:
                print(f"  {row['fecha'].date()}  {str(row['ndoc']):<12}  {str(row['comentarios'])[:45]:<45}  {row['importe']:>12,.2f}")

    if br > 0:
        print('\nBanco sin match:')
        for _, brow in banco.iterrows():
            if bco_col.get(brow['orig_idx']) == ROSA:
                print(f"  {brow['fecha'].date()}  {str(brow['descripcion'])[:45]:<45}  {brow['importe']:>12,.2f}")


# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Uso: python conciliar_brou_usd.py <sap.xlsx> <banco.xls/xlsx>')
        sys.exit(1)

    path_sap   = sys.argv[1]
    path_banco = sys.argv[2]

    sap   = cargar_sap(path_sap)
    banco = cargar_banco(path_banco)

    print(f'SAP: {len(sap)} movimientos | BANCO: {len(banco)} movimientos')

    sap_col, bco_col = conciliar(sap, banco)
    imprimir_resumen(sap, banco, sap_col, bco_col)

    escribir_sap(sap, sap_col, 'CONCILIACION_SAP.xlsx')
    escribir_banco(banco, bco_col, path_banco, 'CONCILIACION_BANCO.xlsx')

    print('\nArchivos generados: CONCILIACION_SAP.xlsx | CONCILIACION_BANCO.xlsx')
